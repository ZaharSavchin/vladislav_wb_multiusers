import asyncio
import json
import re

import aiogram
from aiogram import Router, F
from aiogram.types import Message, InlineKeyboardButton, InlineKeyboardMarkup, FSInputFile
from services.price_monitor import monitoring
from database.database import users_db, users_items, users_max_items, url_images, save_users_items, save_users_db
from aiogram.filters.callback_data import CallbackData
from config_data.config import admin_id
from handlers.currency_handlers import bot

from openpyxl import Workbook


router = Router()


@router.message(F.text.startswith('bot stat'))
async def stat_message(message: Message):
    if message.from_user.id ==admin_id:
        if message.text.endswith('start'):
            await message.answer('цикл запущен')
            # procent = int(message.text.split(' ')[2])
            # print(procent, type(procent))
            procent = 10
            await monitoring(procent)
        elif message.text.endswith("all"):
            answer = []
            counter = 1
            country = {'rub': 0, 'byn': 0, 'kzt': 0, 'kgs': 0, 'uzs': 0, 'usd': 0, 'amd': 0}
            for i in users_db.copy():
                name = users_db[i][0]
                username = users_db[i][1]
                # refs = users_max_items[i]
                if i in users_items:
                    cur = users_items[i][0]
                    country[cur] += 1
                    items = users_items[i][1:]
                    if message.from_user.id == admin_id:
                        answer.append(f"{counter}){name}(@{username}, {i}): {cur}, {items}✅\n")
                    else:
                        answer.append(f"{counter}){name}, @{username}: {cur}\n")

                    counter += 1
                else:
                    if message.from_user.id == admin_id:
                        answer.append(f"{counter}){name}(@{username}, {i})🤷\n")
                    else:
                        answer.append(f"{counter}){name}, @{username}\n")
                    counter += 1

            country_message = f'Россия: {country["rub"]}\n' \
                              f'Беларусь: {country["byn"]}\n' \
                              f'Казахстан: {country["kzt"]}\n' \
                              f'Киргизстан: {country["kgs"]}\n' \
                              f'Узбекистан: {country["uzs"]}\n' \
                              f'Армения: {country["amd"]}\n' \
                              f'В долларах США: {country["usd"]}'

            if len(answer) > 50:
                messages = len(answer) // 50
                counter = 0
                for i in range(messages + 1):
                    stat = ''.join(answer[counter: counter + 50])
                    counter += 50
                    try:
                        await message.answer(f"{stat}")
                    except aiogram.exceptions.TelegramRetryAfter as err:
                        print(err)
                        print('sleep')
                        await asyncio.sleep(200)
                        await message.answer(f"{stat}")
                    await asyncio.sleep(1)
                await message.answer(f'{country_message}')
            else:
                stat = ''.join(answer)
                await message.answer(f"{stat}")
                await message.answer(f'{country_message}')
        else:
            counter = 0
            for i in users_items.copy():
                if len(users_items.copy()[i][1]) > 0:
                    counter += 1
            if message.from_user.id == admin_id:
                await message.answer(f"users: {len(users_db)}\nactive users: {counter}")
            else:
                await message.answer(f"users: {len(users_db)}")


class MaxItemsCallbackFactory(CallbackData, prefix='max_items'):
    user_id: int
    change: str


@router.message(F.text.startswith('change max_items'))
async def change_max_items(message: Message):
    if message.from_user.id == admin_id:
        i = int(message.text.split()[-1])
        name = users_db[i][0]
        username = users_db[i][1]
        refs = users_max_items[i]
        if i in users_items:
            cur = users_items[i][0]
            items = users_items[i][1:]
            answer = f"{name}(@{username}, {i}, {refs}): {cur}, {items}✅\n"
        else:
            answer = f"{name}(@{username}, {i}, {refs})🤷\n"

        button_plus = InlineKeyboardButton(text='+', callback_data=MaxItemsCallbackFactory(user_id=i, change='+').pack())
        button_minus = InlineKeyboardButton(text='-', callback_data=MaxItemsCallbackFactory(user_id=i, change='-').pack())
        markup = InlineKeyboardMarkup(inline_keyboard=[[button_minus, button_plus]])

        await message.answer(text=answer, reply_markup=markup)


@router.message(F.text.startswith('bot curens'))
async def count_cur(message: Message):
    if message.from_user.id == admin_id:
        country = {'rub': 0, 'byn': 0, 'kzt': 0, 'kgs': 0, 'uzs': 0, 'usd': 0, 'amd': 0}
        for i in users_db.copy():
            cur = users_items[i][0]
            country[cur] += 1
        country_message = f'Россия: {country["rub"]}\n' \
                          f'Беларусь: {country["byn"]}\n' \
                          f'Казахстан: {country["kzt"]}\n' \
                          f'Киргизстан: {country["kgs"]}\n' \
                          f'Узбекистан: {country["uzs"]}\n' \
                          f'Армения: {country["amd"]}\n' \
                          f'В долларах США: {country["usd"]}'
        await message.answer(f'{country_message}')


@router.message(F.text.startswith('send ads to users'))
async def send_ads(message: Message):
    if message.from_user.id == admin_id:
        photo_url = 'none'
        text = message.text[message.text.find("}") + 1:]
        if "<" in text or ">" in text:
            text = text.replace(">", "&gt;").replace("<", "&lt;")
        if "{" in message.text and "}" in message.text:
            photo_url = message.text[message.text.find("{") + 1: message.text.find("}")]
        counter = 0
        for user_id, name in users_db.copy().items():
            name = name[0]
            if "<" in name or ">" in name:
                name = name.replace(">", "&gt;").replace("<", "&lt;")
            try:
                if photo_url != 'none':
                    await bot.send_photo(chat_id=user_id,
                                         photo=photo_url,
                                         caption=text)
                else:
                    await bot.send_message(chat_id=user_id,
                                           text=text)
                counter += 1
            except Exception:
                await bot.send_message(chat_id=admin_id, text=f'{user_id}, {name} недоступен')

        await bot.send_message(chat_id=admin_id, text=f'{counter} сообщений доставлено')


def save_dict_to_xlsx(dictionary, filename):
    workbook = Workbook()
    sheet = workbook.active

    keys = list(dictionary.keys())
    for i in range(len(keys)):
        sheet.cell(row=i+1, column=1).value = keys[i]

    values = list(dictionary.values())
    for i in range(len(values)):
        sheet.cell(row=i+1, column=2).value = values[i]

    workbook.save(filename)


@router.message(F.text == 'bot save db')
async def save_db(message: Message):

    # if message.from_user.id == admin_id:
        save_dict_to_xlsx(users_items[admin_id][1], 'wldb.xlsx')

        # with open('users_db.json', 'w', encoding='utf-8-sig') as fl:
        #     json.dump(users_db, fl, indent=4, ensure_ascii=False)
        #
        # with open('users_items.json', 'w', encoding='utf-8-sig') as fl:
        #     json.dump(users_items, fl, indent=4, ensure_ascii=False)
        #
        # with open('users_max_items.json', 'w', encoding='utf-8-sig') as fl:
        #     json.dump(users_max_items, fl, indent=4, ensure_ascii=False)
        #
        # with open('url_images.json', 'w', encoding='utf-8-sig') as fl:
        #     json.dump(url_images, fl, indent=4, ensure_ascii=False)

        # file = FSInputFile('users_db.json')
        # file_1 = FSInputFile('users_items.json')
        file_11 = FSInputFile('wldb.xlsx')
        # file_2 = FSInputFile('users_max_items.json')
        # file_3 = FSInputFile('url_images.json')
        # await message.answer_document(file)
        # await message.answer_document(file_1)
        await message.answer_document(file_11)
        # await message.answer_document(file_2)
        # await message.answer_document(file_3)


@router.message(F.text == 'testprices')
async def shake(message: Message):
    # users_items[message.from_user.id][1]["146894363"] = 0.0
    # users_items[message.from_user.id][1]["146894368"] = 0.0
    # users_items[message.from_user.id][1]["144810560"] = 0.0
    # users_items[message.from_user.id][1]["183887161"] = 4920
    users_items[message.from_user.id][1][159944373] = 0
    users_items[message.from_user.id][1][150249483] = 10080
    await save_users_items()

@router.message(F.text.startswith('delete items'))
async def count_cur(message: Message):
    user_id = message.from_user.id
    if user_id == admin_id:
        counter = 0
        list_items = message.text.split('\n')
        for item in list_items:
            if re.match(r'^\s*\d+\s*$', item):
                if int(item.strip()) in users_items[admin_id][1]:
                    del users_items[admin_id][1][int(item.strip())]
                    counter += 1
        await message.answer(f'{counter} товаров удалено')


